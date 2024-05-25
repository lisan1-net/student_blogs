import re
from functools import lru_cache
from io import BytesIO
from typing import Iterable

from django.core.paginator import Paginator
from django.db.models import Q, Model, Count, Sum
from django.utils.translation import gettext as _
from openpyxl import Workbook
from pyarabic.araby import DIACRITICS

from indexes.models import TextToken, Bigram, Trigram
from indexes.utils import normalize
from main.models import Text, FunctionalWord, Prefix


@lru_cache(64)
def get_diacritics_insensitive_regexp(query: str) -> re.Pattern:
    query = normalize(query)
    return re.compile(''.join(fr'{letter}[{"".join(DIACRITICS)}]*' for letter in query), re.IGNORECASE)


@lru_cache(64)
def find_search_query_position(text: str, query: str, start_index=0) -> tuple[int, int]:
    start = -1
    end = -1
    if query.startswith('"') and query.endswith('"'):
        query = query[1:-1]
        match = re.search(fr'\b{get_diacritics_insensitive_regexp(query).pattern}\b', text[start_index:])
        if match:
            start = match.start() + start_index
            end = match.end() + start_index
    else:
        match = re.search(get_diacritics_insensitive_regexp(query), text[start_index:])
        if match:
            start = match.start() + start_index
            end = match.end() + start_index
    return start, end


@lru_cache(128)
def find_all_search_query_positions(text: str, query: str) -> list[tuple[int, int]]:
    positions = []
    start = 0
    while start != -1:
        start, end = find_search_query_position(text, query, start)
        if start != -1:
            positions.append((start, end))
            start = end
    return positions


def find_search_results(query: str, texts: Iterable) -> list[dict]:
    results = []
    for text in texts:
        for positions in find_all_search_query_positions(text.content, query):
            results.append({'text': text, 'start': positions[0], 'end': positions[1]})
    return results


def build_common_filter_query(cleaned_data: dict) -> (Q, bool):
    advanced_search = False
    filter_query = Q()
    if student_number := cleaned_data['student_number']:
        filter_query &= Q(student_number=student_number)
        advanced_search = True
    if sex := cleaned_data['sex']:
        filter_query &= Q(sex=sex)
        advanced_search = True
    if level := cleaned_data['level']:
        filter_query &= Q(level=level)
        advanced_search = True
    if city := cleaned_data['city']:
        filter_query &= Q(city=city)
        advanced_search = True
    if school := cleaned_data['school']:
        filter_query &= Q(school=school)
        advanced_search = True
    if type_ := cleaned_data['type']:
        filter_query &= Q(type=type_)
        advanced_search = True
    if source_type := cleaned_data['source_type']:
        filter_query &= Q(source_type=source_type)
        advanced_search = True
    if author_name := cleaned_data['author_name']:
        filter_query &= Q(author_name__icontains=author_name)
        advanced_search = True
    if tags := cleaned_data['tags']:
        filter_query &= Q(tags__in=tags)
        advanced_search = True
    return filter_query, advanced_search


@lru_cache(64)
def get_search_results(**cleaned_data):
    filter_query, advanced = build_common_filter_query(cleaned_data)
    if blog := cleaned_data['blog']:
        filter_query &= Q(blog=blog)
        advanced = True
    query = cleaned_data['search_query']
    texts = Text.objects.filter(filter_query).distinct().iterator()
    results = find_search_results(query, texts)
    matched_texts_count = len(set(r['text'] for r in results))
    return results, matched_texts_count, advanced


def get_context(text: str, highlight_start: int, highlight_end: int, surrounding_words):
    text_before = text[:highlight_start]
    highlighted_text = text[highlight_start:highlight_end]
    text_after = text[highlight_end:]
    words_before = text_before.split(' ')
    words_after = text_after.split(' ')
    prefix = Paginator.ELLIPSIS if len(words_before) > surrounding_words else ''
    suffix = Paginator.ELLIPSIS if len(words_after) > surrounding_words else ''
    visible_words_before = words_before[-surrounding_words:]
    visible_words_after = words_after[:surrounding_words]
    return text_before, highlighted_text, text_after, visible_words_before, visible_words_after, prefix, suffix


@lru_cache(32)
def export_search_results(**cleaned_data) -> bytes:
    results, matched_texts_count, advanced = get_search_results(**cleaned_data)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append([
        _('Context'), _('Title'), _('Blog'), _('Student number'), _('Sex'), _('Level'), _('School'), _('City'),
        _('Type'), _('Source')
    ])
    for r in results[:cleaned_data.get('export_count', 500)]:
        text: Text = r['text']
        text_before, highlighted_text, text_after, visible_words_before, visible_words_after, prefix, suffix = get_context(
            text.content, r['start'], r['end'], 5
        )
        ws.append([
            f'{prefix}{" ".join(visible_words_before)}{highlighted_text}{" ".join(visible_words_after)}{suffix}',
            text.title, text.blog.title, text.student_number, text.get_sex_display(), text.level, text.school,
            text.city, text.get_type_display() or text.author_name, text.source_type
        ])
    excel_file = BytesIO()
    wb.save(excel_file)
    return excel_file.getvalue()


@lru_cache(64)
def get_search_paginator_and_counts(**cleaned_data) -> (Paginator, int, bool):
    results, matched_texts_count, advanced = get_search_results(**cleaned_data)
    return Paginator(results, 15), matched_texts_count, advanced


@lru_cache(32)
def get_vocabulary_results(**cleaned_data):
    filter_query, advanced = build_common_filter_query(cleaned_data)
    if blog_id := cleaned_data['blog']:
        filter_query &= Q(blog_id=blog_id)
    texts = Text.objects.filter(filter_query).distinct()
    words = TextToken.objects.filter(text__in=texts).values('token__content')
    include_functional_words = cleaned_data['include_functional_words']
    contains = cleaned_data['contains']
    partial_contains = cleaned_data['partial_contains']
    not_contains = cleaned_data['not_contains']
    partial_not_contains = cleaned_data['partial_not_contains']
    if contains:
        if partial_contains:
            words = words.filter(token__content__icontains=contains)
        else:
            words = words.filter(token__content=contains)
    if not_contains:
        if partial_not_contains:
            words = words.exclude(token__content__icontains=not_contains)
        else:
            words = words.exclude(token__content=not_contains)
    if not include_functional_words:
        words = words.exclude(token__content__in=FunctionalWord.objects.values_list('content', flat=True))
    word_frequencies = words.annotate(frequency=Count('token__content')).order_by('-frequency').values_list(
        'token__content', 'frequency'
    )
    return word_frequencies


@lru_cache(32)
def export_vocabulary_results(**cleaned_data):
    word_frequencies = get_vocabulary_results(**cleaned_data)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append([_('Word'), _('Frequency')])
    for word, frequency in word_frequencies[:cleaned_data.get('export_count', 500)]:
        ws.append([word, frequency])
    excel_file = BytesIO()
    wb.save(excel_file)
    return excel_file.getvalue()


@lru_cache(64)
def get_vocabulary_paginator(**cleaned_data) -> Paginator:
    word_frequencies = get_vocabulary_results(**cleaned_data)
    return Paginator(word_frequencies, 60)


@lru_cache(64)
def get_ngrams_paginator(**cleaned_data) -> Paginator:
    blog_id = cleaned_data['blog']
    filter_query = Q(blog_id=blog_id)
    q, advanced = build_common_filter_query(cleaned_data)
    filter_query &= q
    texts = Text.objects.filter(filter_query).distinct()
    ngram_type = cleaned_data['ngram_type']
    ngrams = []
    contains = cleaned_data['contains']
    partial_contains = cleaned_data['partial_contains']
    not_contains = cleaned_data['not_contains']
    partial_not_contains = cleaned_data['partial_not_contains']
    match ngram_type:
        case 'bigram':
            ngrams = Bigram.objects.filter(text__in=texts).prefetch_related('first_token', 'second_token').values(
                'first_token__content', 'second_token__content'
            )
            if contains:
                if partial_contains:
                    ngrams = ngrams.filter(
                        Q(first_token__content__icontains=contains) | Q(second_token__content__icontains=contains)
                    )
                else:
                    ngrams = ngrams.filter(
                        Q(first_token__content=contains) | Q(second_token__content=contains)
                    )
            if not_contains:
                if partial_not_contains:
                    ngrams = ngrams.exclude(
                        Q(first_token__content__icontains=not_contains) |
                        Q(second_token__content__icontains=not_contains)
                    )
                else:
                    ngrams = ngrams.exclude(
                        Q(first_token__content=not_contains) | Q(second_token__content=not_contains)
                    )
            ngrams = ngrams.annotate(frequency=Sum('frequency')).order_by('-frequency').values_list(
                'first_token__content', 'second_token__content', 'frequency'
            )
        case 'trigram':
            ngrams = Trigram.objects.filter(text__in=texts).prefetch_related(
                'first_token', 'second_token', 'third_token'
            ).values(
                'first_token__content', 'second_token__content', 'third_token__content'
            )
            if contains:
                if partial_contains:
                    ngrams = ngrams.filter(
                        Q(first_token__content__icontains=contains) | Q(second_token__content__icontains=contains) |
                        Q(third_token__content__icontains=contains)
                    )
                else:
                    ngrams = ngrams.filter(
                        Q(first_token__content=contains) | Q(second_token__content=contains) |
                        Q(third_token__content=contains)
                    )
            if not_contains:
                if partial_not_contains:
                    ngrams = ngrams.exclude(
                        Q(first_token__content__icontains=not_contains) |
                        Q(second_token__content__icontains=not_contains) |
                        Q(third_token__content__icontains=not_contains)
                    )
                else:
                    ngrams = ngrams.exclude(
                        Q(first_token__content=not_contains) | Q(second_token__content=not_contains) |
                        Q(third_token__content=not_contains)
                    )
            ngrams = ngrams.annotate(frequency=Sum('frequency')).order_by('-frequency').values_list(
                'first_token__content', 'second_token__content', 'third_token__content', 'frequency'
            )
    return Paginator(ngrams, 60)


@lru_cache(64)
def get_surrounding_words_frequencies_paginator(**cleaned_data) -> (Paginator, bool):
    filter_query, advanced = build_common_filter_query(cleaned_data)
    if blog_id := cleaned_data['blog']:
        filter_query &= Q(blog_id=blog_id)
    texts = Text.objects.filter(filter_query).distinct()
    query = normalize(cleaned_data['search_query'])
    fully_indexed = not texts.filter(bigrams_indexed=False).exists()
    filters = dict(text__in=texts)
    if cleaned_data['position'] == 'P':
        if cleaned_data['partial_search']:
            key = 'second_token__content__icontains'
        else:
            key = 'second_token__content'
        filters[key] = query
    elif cleaned_data['position'] == 'N':
        if cleaned_data['partial_search']:
            key = 'first_token__content__icontains'
        else:
            key = 'first_token__content'
        filters[key] = query
    frequencies = Bigram.objects.filter(**filters).distinct().prefetch_related(
        'first_token', 'second_token'
    ).values('second_token__content').annotate(frequency=Sum('frequency')).order_by('-frequency').values_list(
        'first_token__content', 'second_token__content', 'frequency'
    )
    return Paginator(frequencies, 60), fully_indexed


@lru_cache(128)
def get_possible_derivations(word: str):
    possible_derivations = set()
    for prefix in Prefix.objects.all():
        possible_derivations.add(f'{prefix}{word}')
        for suffix in prefix.suffixes.all():
            possible_derivations.add(f'{word}{suffix}')
            possible_derivations.add(f'{prefix}{word}{suffix}')
    return possible_derivations


@lru_cache(64)
def get_derivation_frequencies_paginator(**cleaned_data) -> (Paginator, bool):
    filter_query, advanced = build_common_filter_query(cleaned_data)
    if blog_id := cleaned_data['blog']:
        filter_query &= Q(blog_id=blog_id)
    texts = Text.objects.filter(filter_query).distinct()
    word = normalize(cleaned_data['search_query'])
    possible_derivations = get_possible_derivations(word)
    frequencies = TextToken.objects.filter(text__in=texts, token__content__in=possible_derivations).values(
        'token__content'
    ).annotate(frequency=Count('token__content')).order_by('-frequency').values_list('token__content', 'frequency')
    fully_indexed = not texts.filter(words_indexed=False).exists()
    return Paginator(frequencies, 60), fully_indexed


def clean_form_data(form_data: dict) -> dict:
    form_data = form_data.copy()
    for k, v in form_data.items():
        if isinstance(v, Iterable) and not isinstance(v, str):
            form_data[k] = tuple(v)
        elif isinstance(v, Model):
            form_data[k] = v.pk
    return form_data
